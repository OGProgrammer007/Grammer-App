import pygame
import openpyxl
import os
import time
import numpy as np
from PIL import Image
import streamlit as st

# Excel file path
EXCEL_FILE = "Scoreboard_project/scores_with_avatars.xlsx"

# Folder where avatar images are stored in the repository
AVATAR_FOLDER = "Scoreboard_project/"

# Funky music file path
MUSIC_FILE = "Scoreboard_project/funky_music.mp3"

# Colors
COLORS = {
    "gold": "#FFD700",
    "silver": "#C0C0C0",
    "bronze": "#CD7F32",
    "blue": "#6495ED",
    "white": "#FFFFFF"
}

# Initialize pygame
pygame.init()

# Screen dimensions
WIDTH, HEIGHT = 800, 700
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Leaderboard with Top 5 and Scrollable List")

# Fonts
TITLE_FONT = pygame.font.Font(None, 72)
ENTRY_FONT = pygame.font.Font(None, 48)
SCORE_FONT = pygame.font.Font(None, 36)

def load_scores():
    """Load player data from the Excel file."""
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: File '{EXCEL_FILE}' not found.")
        return {}

    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    scores = {}

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        name, score, avatar = row[0].value, row[1].value, row[2].value
        if name and score and avatar:
            scores[name] = {"score": int(score), "avatar": avatar}

    workbook.close()
    return scores

def load_avatar(path):
    """Load and resize avatars."""
    if not os.path.exists(path):
        print(f"Warning: Avatar '{path}' not found.")
        return None
    try:
        avatar = pygame.image.load(path)
        return pygame.transform.scale(avatar, (100, 100))
    except pygame.error:
        print(f"Error loading image '{path}'")
        return None

# Load scores and avatars
players = load_scores()
avatars = {name: load_avatar(os.path.join(AVATAR_FOLDER, data["avatar"])) for name, data in players.items()}

def rotate_image(image, angle):
    """Rotate image side-to-side effect."""
    return image.rotate(angle, expand=True)

def draw_leaderboard(angle):
    """Draw the leaderboard with top 5 players and scrollable remaining players."""
    screen.fill((0, 0, 0))

    # Title
    title = TITLE_FONT.render("Leaderboard", True, COLORS["gold"])
    screen.blit(title, ((WIDTH - title.get_width()) // 2, 20))

    if not players:
        msg = SCORE_FONT.render("No data available", True, COLORS["white"])
        screen.blit(msg, ((WIDTH - msg.get_width()) // 2, HEIGHT // 2))
        pygame.display.update()
        return

    sorted_players = sorted(players.items(), key=lambda x: x[1]["score"], reverse=True)

    # Top 5 players
    y_start = 100
    for i, (player, data) in enumerate(sorted_players[:5]):
        player_name = f"{i + 1}. {player}"
        player_score = f"{data['score']} points"

        player_text = ENTRY_FONT.render(player_name, True, COLORS["gold"] if i == 0 else COLORS["silver"] if i == 1 else COLORS["bronze"] if i == 2 else COLORS["white"])
        score_text = SCORE_FONT.render(player_score, True, COLORS["blue"])

        screen.blit(player_text, (200, y_start + i * 120))
        screen.blit(score_text, (580, y_start + i * 120))

        # Draw rotating avatar
        avatar = avatars.get(player)
        if avatar:
            rotated_avatar = rotate_image(avatar, np.sin(angle + i) * 30)
            avatar_rect = rotated_avatar.get_rect(center=(100, y_start + i * 100 + 65))
            screen.blit(rotated_avatar, avatar_rect.topleft)

    # Scrollable section for remaining players
    with st.expander("View all players"):
        for i, (player, data) in enumerate(sorted_players[5:]):
            col1, col2 = st.columns([1, 3])

            with col1:
                avatar_path = os.path.join(AVATAR_FOLDER, data["avatar"])
                avatar = load_avatar(avatar_path)
                if avatar:
                    # Apply rotation
                    angle = np.sin(time.time()) * 30  # Continuous rotation effect
                    rotated_avatar = rotate_image(avatar, angle)
                    st.image(rotated_avatar, width=100)

            with col2:
                st.markdown(f"<h3 style='color:{COLORS['white']}'>{i + 1}. {player}</h3>", unsafe_allow_html=True)
                st.markdown(f"**{data['score']} points**", unsafe_allow_html=True)

# Run the pygame window to show the leaderboard
running = True
angle = 0  # Angle for rotating the avatars

while running:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

    angle += 0.1

    draw_leaderboard(angle)

    pygame.display.update()  # Refresh/update the scores
    pygame.time.delay(50)  # Delay to control the speed

pygame.quit()
