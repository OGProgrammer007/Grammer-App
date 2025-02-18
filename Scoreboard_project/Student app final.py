import pygame
import openpyxl
import math
import os

pygame.init()

# Screen dimensions
WIDTH, HEIGHT = 800, 700
screen = pygame.display.set_mode((WIDTH, HEIGHT))
pygame.display.set_caption("Leaderboard with Top 5 and Scrollable List")

# Colors
WHITE = (255, 255, 255)
BLACK = (0, 0, 0)
GOLD = (255, 215, 0)
SILVER = (192, 192, 192)
BRONZE = (205, 127, 50)
BLUE = (100, 149, 237)
GRAY = (200, 200, 200)

# Fonts
TITLE_FONT = pygame.font.Font(None, 72)
ENTRY_FONT = pygame.font.Font(None, 48)
SCORE_FONT = pygame.font.Font(None, 36)
SMALL_FONT = pygame.font.Font(None, 28)

# Excel file for scores
EXCEL_FILE = "scores_with_avatars.xlsx"

def load_scores():
    """Load player data from the Excel file."""
    if not os.path.exists(EXCEL_FILE):
        print(f"Error: File '{EXCEL_FILE}' not found.")
        return {}

    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active
    scores = {}
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
        name, score, avatar = row[0].value, row[1].value, row[2].value
        if name and score and avatar:
            scores[name] = {"score": int(score), "avatar": avatar}
    
    workbook.close()
    return scores

def load_avatar(path):
    """Load and resize avatars."""
    if not os.path.exists(path):
        print(f"Warning: Avatar '{path}' not found.")
        return None
    try:
        avatar = pygame.image.load(path)
        return pygame.transform.scale(avatar, (100, 100))
    except pygame.error:
        print(f"Error loading image '{path}'")
        return None

# Load scores and avatars
players = load_scores()
avatars = {name: load_avatar(data["avatar"]) for name, data in players.items()}

def draw_leaderboard(angle):
    """Draw the leaderboard with top 5 players and scrollable remaining players."""
    screen.fill(BLACK)

    # Title
    title = TITLE_FONT.render("Leaderboard", True, GOLD)
    screen.blit(title, ((WIDTH - title.get_width()) // 2, 20))

    if not players:
        msg = SCORE_FONT.render("No data available", True, WHITE)
        screen.blit(msg, ((WIDTH - msg.get_width()) // 2, HEIGHT // 2))
        pygame.display.update()
        return

    sorted_players = sorted(players.items(), key=lambda x: x[1]["score"], reverse=True)

    # Top 5 players
    y_start = 100
    colors = [GOLD, BRONZE] + [WHITE] * max(0, (len(sorted_players) - 1))

    for i, (player, data) in enumerate(sorted_players[:5]):
        player_name = f"{i + 1}. {player}"
        player_score = f"{data['score']} points"

        player_text = ENTRY_FONT.render(player_name, True, colors[i])
        score_text = SCORE_FONT.render(player_score, True, BLUE)

        screen.blit(player_text, (200, y_start + i * 120))
        screen.blit(score_text, (580, y_start + i * 120))

        # Draw tilting avatar
        avatar = avatars.get(player)
        if avatar:
            rotated_avatar = pygame.transform.rotate(avatar, math.sin(angle + i) * 30)
            avatar_rect = rotated_avatar.get_rect(center=(100, y_start + i * 100 + 65))
            screen.blit(rotated_avatar, avatar_rect.topleft)

running = True
angle = 0  # Angle for tilting the avatars

while running:
    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False

    angle += 0.1

    draw_leaderboard(angle)

    pygame.display.update() #refresh/update the scores
    pygame.time.delay(50)

pygame.quit()
